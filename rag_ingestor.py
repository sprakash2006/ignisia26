import os
import re
import csv
import email
import email.policy
from email.utils import parsedate_to_datetime


class FileIngestor:
    def process_file(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        filename = os.path.basename(file_path)

        all_chunks = []

        if ext == ".pdf":
            all_chunks = self._process_pdf(file_path)
        elif ext == ".docx":
            all_chunks = self._process_docx(file_path)
        elif ext == ".xlsx":
            all_chunks = self._process_excel(file_path)
        elif ext == ".csv":
            all_chunks = self._process_csv(file_path)
        elif ext == ".txt":
            text = self._extract_text_from_txt(file_path)
            all_chunks = [{"text": chunk, "page": 1, "line": i + 1} for i, chunk in enumerate(self._split_into_chunks(text))]
        elif ext == ".eml":
            all_chunks = self._process_email(file_path)
        else:
            print(f"[WARN] Unsupported file type: {ext}")
            return [], filename

        print(f"[INFO] Extracted {len(all_chunks)} chunks from {file_path}")
        return all_chunks, filename

    # ── PDF processing (pdfplumber) ─────────────────────────────────

    def _process_pdf(self, file_path):
        try:
            import pdfplumber
            all_chunks = []
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text()
                    if not text or not text.strip():
                        continue
                    # Split page text into chunks
                    page_chunks = self._split_into_chunks(text, max_length=1000)
                    for i, chunk in enumerate(page_chunks):
                        all_chunks.append({
                            "text": chunk,
                            "page": page_num,
                            "line": i + 1,
                        })
            return all_chunks
        except Exception as e:
            print(f"[ERROR] PDF extraction failed: {e}")
            return []

    # ── DOCX processing (python-docx) ──────────────────────────────

    def _process_docx(self, file_path):
        try:
            from docx import Document
            doc = Document(file_path)
            full_text = []
            current_section = ""

            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    continue
                # Detect headings for section tracking
                if para.style and para.style.name.startswith("Heading"):
                    current_section = text
                full_text.append(text)

            combined = "\n\n".join(full_text)
            chunks = self._split_into_chunks(combined, max_length=1000)
            return [{"text": chunk, "page": 1, "line": i + 1, "section": current_section} for i, chunk in enumerate(chunks)]
        except Exception as e:
            print(f"[ERROR] DOCX extraction failed: {e}")
            return []

    # ── Text helpers ────────────────────────────────────────────────

    def _extract_text_from_txt(self, file_path):
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    def _split_into_chunks(self, text, max_length=500):
        if not text.strip():
            return []

        # Split by double newline first (paragraphs)
        paragraphs = text.split("\n\n")
        chunks, current_chunk = [], ""

        for para in paragraphs:
            para = para.strip()
            if not para:
                continue

            # If a paragraph is still too long, split it by single newline
            if len(para) > max_length:
                sub_paras = para.split("\n")
                for sub in sub_paras:
                    sub = sub.strip()
                    if not sub:
                        continue
                    if len(current_chunk) + len(sub) < max_length:
                        current_chunk += sub + " "
                    else:
                        if current_chunk.strip():
                            chunks.append(current_chunk.strip())
                        current_chunk = sub + " "
            else:
                if len(current_chunk) + len(para) < max_length:
                    current_chunk += para + "\n\n"
                else:
                    if current_chunk.strip():
                        chunks.append(current_chunk.strip())
                    current_chunk = para + "\n\n"

        if current_chunk.strip():
            chunks.append(current_chunk.strip())
        return chunks

    # ── Excel processing ────────────────────────────────────────────

    def _process_excel(self, file_path):
        """Process .xlsx files — each row becomes a chunk with column headers as keys."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            all_chunks = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                # First row is header
                headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
                for row_idx, row in enumerate(rows[1:], start=2):
                    parts = []
                    for header, val in zip(headers, row):
                        cell_val = str(val).strip() if val is not None else ""
                        if cell_val:
                            parts.append(f"{header}: {cell_val}")
                    if parts:
                        text = f"[Sheet: {sheet_name}, Row: {row_idx}] " + " | ".join(parts)
                        all_chunks.append({
                            "text": text,
                            "page": 1,
                            "line": row_idx,
                            "section": sheet_name,
                        })
            wb.close()
            return all_chunks
        except Exception as e:
            print(f"[ERROR] Excel processing failed: {e}")
            return []

    # ── CSV processing ──────────────────────────────────────────────

    def _process_csv(self, file_path):
        """Process .csv files — each row becomes a chunk with column headers as keys."""
        try:
            all_chunks = []
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if not rows:
                return []
            headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(rows[0])]
            for row_idx, row in enumerate(rows[1:], start=2):
                parts = []
                for header, val in zip(headers, row):
                    cell_val = str(val).strip() if val else ""
                    if cell_val:
                        parts.append(f"{header}: {cell_val}")
                if parts:
                    text = f"[Row: {row_idx}] " + " | ".join(parts)
                    all_chunks.append({
                        "text": text,
                        "page": 1,
                        "line": row_idx,
                        "section": "CSV",
                    })
            return all_chunks
        except Exception as e:
            print(f"[ERROR] CSV processing failed: {e}")
            return []

    # ── Email (.eml) processing ─────────────────────────────────────

    def _process_email(self, file_path):
        """Parse a .eml file and return chunks with email metadata."""
        try:
            with open(file_path, "rb") as f:
                msg = email.message_from_binary_file(f, policy=email.policy.default)

            # Extract header fields
            sender = str(msg.get("From", "Unknown"))
            subject = str(msg.get("Subject", "(No Subject)"))
            to = str(msg.get("To", "Unknown"))

            # Parse actual email date
            email_date = None
            raw_date = msg.get("Date")
            if raw_date:
                try:
                    dt = parsedate_to_datetime(str(raw_date))
                    email_date = dt.strftime("%Y-%m-%d")
                except Exception:
                    email_date = None

            # Extract plain-text body
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type()
                    if ctype == "text/plain":
                        payload = part.get_content()
                        if isinstance(payload, str):
                            body += payload
            else:
                payload = msg.get_content()
                if isinstance(payload, str):
                    body = payload

            if not body.strip():
                print(f"[WARN] No text body found in email: {file_path}")
                return []

            # Split long email threads on common reply/forward markers
            thread_parts = re.split(
                r'\n\s*(?:On .+ wrote:|---------- Forwarded message ----------|From:.*Sent:)',
                body
            )

            header_line = (
                f"[Email from: {sender}, To: {to}, "
                f"Subject: {subject}, Date: {email_date or 'unknown'}]"
            )

            all_chunks = []
            for i, part in enumerate(thread_parts):
                part = part.strip()
                if not part or len(part) < 20:
                    continue

                if len(part) > 500:
                    sub_chunks = self._split_into_chunks(part, max_length=500)
                    for j, sc in enumerate(sub_chunks):
                        all_chunks.append({
                            "text": f"{header_line}\n{sc}",
                            "page": 1,
                            "line": i + 1,
                            "section": f"Thread part {i + 1}",
                            "source_date": email_date,
                        })
                else:
                    all_chunks.append({
                        "text": f"{header_line}\n{part}",
                        "page": 1,
                        "line": i + 1,
                        "section": f"Thread part {i + 1}" if len(thread_parts) > 1 else "Email body",
                        "source_date": email_date,
                    })

            return all_chunks
        except Exception as e:
            print(f"[ERROR] Email processing failed: {e}")
            return []
